import pdfplumber
import openpyxl
import os
import logging
import subprocess

logger = logging.getLogger('converter')

def convert_pdf_to_excel(input_pdf_path, output_excel_path, mode='pdfplumber'):
    """
    Converts tables from a PDF file to an Excel file.
    Each table in the PDF will be a sheet in the Excel file.

    Args:
        input_pdf_path (str): Path to the input PDF file.
        output_excel_path (str): Path to save the output Excel file.
        mode (str): Conversion mode - 'pdfplumber' (default) or 'libreoffice'

    Returns:
        tuple: (success: bool, actual_output_path: str or None, error_message: str or None)
    """
    try:
        logger.info(f"Starting PDF to Excel conversion using {mode} mode: {input_pdf_path} -> {output_excel_path}")

        if not os.path.exists(input_pdf_path):
            return False, None, "Input PDF file not found."

        if mode == 'libreoffice':
            return _convert_pdf_to_excel_libreoffice(input_pdf_path, output_excel_path)
        else:  # Default to pdfplumber mode
            return _convert_pdf_to_excel_pdfplumber(input_pdf_path, output_excel_path)

    except Exception as e:
        logger.error(f"Error during PDF to Excel conversion for '{input_pdf_path}' using {mode}: {e}", exc_info=True)
        return False, None, f"PDF转Excel失败: {str(e)}"

def _convert_pdf_to_excel_pdfplumber(input_pdf_path, output_excel_path):
    """Convert PDF to Excel using pdfplumber."""
    try:
        workbook = openpyxl.Workbook()
        # Remove default sheet created by openpyxl
        if "Sheet" in workbook.sheetnames:
            default_sheet = workbook["Sheet"]
            workbook.remove(default_sheet)
        
        has_tables = False
        with pdfplumber.open(input_pdf_path) as pdf:
            if not pdf.pages:
                return False, None, "PDF has no pages."

            for i, page in enumerate(pdf.pages):
                # Extract tables from the current page
                tables_on_page = page.extract_tables()
                
                if tables_on_page:
                    has_tables = True
                    for table_idx, table_data in enumerate(tables_on_page):
                        # Create a new sheet for each table
                        sheet_title = f"Page{i+1}_Table{table_idx+1}"
                        # Truncate sheet title if too long (Excel limit is 31 chars)
                        if len(sheet_title) > 31:
                            sheet_title = sheet_title[:31]
                        
                        # Ensure sheet title is unique if truncated or many tables
                        original_sheet_title = sheet_title
                        counter = 1
                        while sheet_title in workbook.sheetnames:
                            suffix = f"_{counter}"
                            if len(original_sheet_title) + len(suffix) > 31:
                                sheet_title = original_sheet_title[:31-len(suffix)] + suffix
                            else:
                                sheet_title = original_sheet_title + suffix
                            counter += 1
                            if counter > 100: # Safety break
                                return False, None, "Too many tables with similar names, cannot create unique sheet names."

                        sheet = workbook.create_sheet(title=sheet_title)
                        
                        for row_data in table_data:
                            # pdfplumber might return None for cells if they are merged or empty in a complex way
                            # Replace None with empty string for openpyxl
                            cleaned_row = [str(cell) if cell is not None else "" for cell in row_data]
                            sheet.append(cleaned_row)
                        logger.info(f"Added table to sheet: {sheet_title}")
                else:
                    logger.info(f"No tables found on page {i+1}")

        if not has_tables:
            logger.warning(f"No tables found in the PDF using pdfplumber: {input_pdf_path}")
            return False, None, "未在PDF中找到可供转换的表格。"

        workbook.save(output_excel_path)
        logger.info(f"Successfully converted PDF to Excel using pdfplumber: {output_excel_path}")
        return True, output_excel_path, None

    except Exception as e:
        logger.error(f"Error during pdfplumber PDF to Excel conversion: {e}", exc_info=True)
        return False, None, f"pdfplumber转换失败: {str(e)}"

def _convert_pdf_to_excel_libreoffice(input_pdf_path, output_excel_path):
    """Convert PDF to Excel using LibreOffice (simplified approach based on PDF to PPT success pattern)."""
    try:
        output_dir = os.path.dirname(output_excel_path)
        os.makedirs(output_dir, exist_ok=True)
        
        logger.info(f"Using LibreOffice to convert {input_pdf_path} to {output_excel_path}")
        
        # Calculate expected output path (LibreOffice uses input file's basename)
        input_basename = os.path.splitext(os.path.basename(input_pdf_path))[0]
        temp_soffice_output_path = os.path.join(output_dir, f"{input_basename}.xlsx")
        
        try:
            # Use single approach similar to successful PDF to PPT implementation
            cmd = [
                'soffice', 
                '--headless',
                '--infilter=calc_pdf_import',
                '--convert-to', 'xlsx',
                '--outdir', output_dir,
                input_pdf_path
            ]
            logger.info(f"Executing LibreOffice command: {' '.join(cmd)}")
            process = subprocess.run(cmd, capture_output=True, text=True, timeout=120)

            if process.returncode == 0:
                logger.info(f"LibreOffice process completed successfully for {input_pdf_path}.")
                
                # Check if soffice created the file with expected name
                if os.path.exists(temp_soffice_output_path):
                    # Rename to desired output path if different
                    if temp_soffice_output_path != output_excel_path:
                        if os.path.exists(output_excel_path):
                            logger.warning(f"Target file {output_excel_path} already exists. Overwriting for LibreOffice conversion.")
                            os.remove(output_excel_path)
                        os.rename(temp_soffice_output_path, output_excel_path)
                        logger.info(f"Renamed LibreOffice output from {temp_soffice_output_path} to {output_excel_path}")
                    
                    if os.path.exists(output_excel_path):
                        logger.info(f"LibreOffice conversion successful: {output_excel_path}")
                        return True, output_excel_path, "LibreOffice conversion successful."
                    else:
                        logger.error(f"LibreOffice conversion error: Expected output file {output_excel_path} not found after potential rename.")
                        return False, None, "LibreOffice conversion failed: Output file not found after rename."
                else:
                    logger.error(f"LibreOffice conversion error: Expected intermediate output {temp_soffice_output_path} not found. Soffice stdout: {process.stdout}, stderr: {process.stderr}")
                    
                    # Provide specific guidance for "no export filter" error
                    if "no export filter" in process.stderr.lower():
                        error_msg = ("LibreOffice缺少PDF转Excel的导出过滤器。\n\n"
                                   "💡 推荐解决方案：\n"
                                   "1. 在页面上选择'使用pdfplumber(默认)'模式\n"
                                   "2. pdfplumber专门优化PDF表格提取，效果通常更好\n"
                                   "3. 重新上传文件并转换即可\n\n"
                                   "📋 技术详情：LibreOffice安装缺少calc_pdf_import过滤器")
                        return False, None, error_msg
                    else:
                        return False, None, f"LibreOffice conversion failed: Soffice did not produce expected output file. Details: {process.stderr[:200]}"
            else:
                logger.error(f"LibreOffice conversion failed for {input_pdf_path}. Return code: {process.returncode}")
                logger.error(f"LibreOffice stdout: {process.stdout}")
                logger.error(f"LibreOffice stderr: {process.stderr}")
                return False, None, f"LibreOffice conversion failed. Error: {process.stderr[:200]}"

        except FileNotFoundError:
            logger.error("LibreOffice (soffice) command not found. Ensure LibreOffice is installed and in PATH.")
            return False, None, "LibreOffice (soffice) command not found. Please install LibreOffice and add it to your system PATH."
        except subprocess.TimeoutExpired:
            logger.error(f"LibreOffice conversion timed out for {input_pdf_path}.")
            return False, None, "LibreOffice conversion timed out."
        except Exception as e_lo:
            logger.error(f"Error during LibreOffice conversion for {input_pdf_path}: {e_lo}", exc_info=True)
            # Clean up partial files
            if os.path.exists(temp_soffice_output_path) and temp_soffice_output_path != output_excel_path: 
                try: os.remove(temp_soffice_output_path) 
                except: pass
            if os.path.exists(output_excel_path):
                 try: os.remove(output_excel_path)
                 except: pass
            return False, None, f"LibreOffice conversion error: {str(e_lo)}"
        
    except Exception as e:
        error_msg = ("LibreOffice PDF转Excel功能在当前环境中不可用。\n\n"
                    "💡 解决方案：\n"
                    "1. 在页面上选择'使用pdfplumber(默认)'模式\n"
                    "2. pdfplumber模式专门优化了PDF表格提取，通常效果更好\n"
                    "3. 重新上传文件并转换即可")
        logger.error(error_msg, exc_info=True)
        return False, None, error_msg

if __name__ == '__main__':
    # Create dummy PDF and test (requires reportlab for dummy creation)
    # This part is for standalone testing and might require additional setup/libraries
    # For now, just provide a simple test call structure

    # To test this, you'd need a sample PDF file.
    # Example:
    # success, out_path, err = convert_pdf_to_excel("sample.pdf", "output.xlsx")
    # if success:
    #     print(f"Conversion successful: {out_path}")
    # else:
    #     print(f"Conversion failed: {err}")
    pass 