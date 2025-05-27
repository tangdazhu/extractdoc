import subprocess
import os
import logging
from pathlib import Path
import tempfile
import shutil
# Import the generic LibreOffice converter
from .libreoffice_converter import convert_to_pdf as lo_convert_to_pdf

logger = logging.getLogger('converter')

def register_chinese_fonts():
    """注册中文字体以支持中文显示"""
    try:
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        from reportlab.lib.fonts import addMapping
        
        # 尝试注册常见的中文字体
        font_paths = [
            'C:/Windows/Fonts/simsun.ttc',  # 宋体
            'C:/Windows/Fonts/simhei.ttf',  # 黑体
            'C:/Windows/Fonts/simkai.ttf',  # 楷体
            'C:/Windows/Fonts/msyh.ttc',    # 微软雅黑
            # For Linux, common paths could be:
            # '/usr/share/fonts/truetype/wqy/wqy-microhei.ttc',
            # '/usr/share/fonts/truetype/noto/NotoSansCJK-Regular.ttc',
        ]
        
        # 为Linux环境调整字体名称和添加映射的逻辑可能需要更通用
        # 例如，不仅仅基于文件名后缀，可能需要检查字体元数据或使用更标准的名称

        for font_path in font_paths:
            if os.path.exists(font_path):
                try:
                    font_name_map = {
                        'simsun.ttc': ('SimSun', 0),
                        'simhei.ttf': ('SimHei', None),
                        'simkai.ttf': ('SimKai', None),
                        'msyh.ttc': ('Microsoft-YaHei', 0),
                        'wqy-microhei.ttc': ('WenQuanYiMicroHei', None), # Example for Linux
                        'NotoSansCJK-Regular.ttc': ('NotoSansCJK', None) # Example for Linux
                    }
                    
                    base_font_file = os.path.basename(font_path)
                    if base_font_file in font_name_map:
                        font_register_name, subfont_idx = font_name_map[base_font_file]
                        if subfont_idx is not None:
                            pdfmetrics.registerFont(TTFont(font_register_name, font_path, subfontIndex=subfont_idx))
                        else:
                            pdfmetrics.registerFont(TTFont(font_register_name, font_path))
                        
                        # Basic mapping, might need refinement for bold/italic if separate font files are not used
                        addMapping(font_register_name, 0, 0, font_register_name) # normal
                        addMapping(font_register_name, 1, 0, font_register_name) # bold (map to normal if no bold variant registered)
                        addMapping(font_register_name, 0, 1, font_register_name) # italic (map to normal)
                        addMapping(font_register_name, 1, 1, font_register_name) # bold-italic (map to normal)

                        logger.info(f"Successfully registered font: {font_register_name} from {font_path}")
                        return font_register_name # Return the first successfully registered font
                except Exception as e:
                    logger.warning(f"Failed to register font {font_path}: {e}")
                    continue
        
        logger.warning("No suitable Chinese/CJK fonts found or registered from predefined paths, using default Helvetica.")
        return 'Helvetica'
        
    except ImportError:
        logger.warning("ReportLab not available for font registration, using default Helvetica.")
        return 'Helvetica'

def convert_excel_to_pdf_openpyxl(input_path, output_path):
    """
    使用openpyxl和reportlab创建带工作表标题的优化PDF，支持中文和超宽表格
    
    Args:
        input_path: 输入的Excel文件路径(.xls/.xlsx)
        output_path: 输出的PDF文件路径
    
    Returns:
        tuple: (success: bool, actual_output_path: str or None, error_message: str or None)
    """
    try:
        from openpyxl import load_workbook
        from reportlab.lib.pagesizes import A4, A3, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors
        from reportlab.lib.units import inch, cm
        
        chinese_font = register_chinese_fonts()
        
        wb = load_workbook(input_path, data_only=True)
        page_size = landscape(A3)
        doc = SimpleDocTemplate(
            output_path,
            pagesize=page_size,
            rightMargin=0.3*cm,
            leftMargin=0.3*cm,
            topMargin=0.5*cm,
            bottomMargin=0.3*cm
        )
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontName=chinese_font,
            fontSize=18,
            spaceAfter=20,
            alignment=1,
            textColor=colors.darkblue
        )
        story = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            title = Paragraph(f"<b>{sheet_name}</b>", title_style)
            story.append(title)
            story.append(Spacer(1, 10))
            data = []
            max_col = 0
            for row in ws.iter_rows(values_only=True):
                if any(cell is not None for cell in row):
                    row_data = [str(cell).strip() if cell is not None else '' for cell in row]
                    data.append(row_data)
                    max_col = max(max_col, len(row_data))
            if data:
                for row in data:
                    while len(row) < max_col:
                        row.append('')
                available_width = page_size[0] - doc.leftMargin - doc.rightMargin
                col_width = available_width / max_col if max_col > 0 else available_width / 10
                min_col_width = 0.8*cm
                max_col_width = 4*cm
                col_width = max(min_col_width, min(col_width, max_col_width))
                table = Table(data, colWidths=[col_width] * max_col)
                table_style = TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ('FONTNAME', (0, 0), (-1, 0), chinese_font),
                    ('FONTSIZE', (0, 0), (-1, 0), 9),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                    ('FONTNAME', (0, 1), (-1, -1), chinese_font),
                    ('FONTSIZE', (0, 1), (-1, -1), 8),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
                ])
                table.setStyle(table_style)
                story.append(table)
                story.append(Spacer(1, 20))
                logger.info(f"Processed worksheet '{sheet_name}' with {len(data)} rows and {max_col} columns, col_width: {col_width}")
        doc.build(story)
        if os.path.exists(output_path):
            logger.info(f"OpenPyXL转换成功: {output_path}")
            return True, output_path, None
        else:
            error_msg = "OpenPyXL转换完成，但未找到输出文件"
            logger.error(error_msg)
            return False, None, error_msg
    except ImportError as e:
        error_msg = f"所需库未安装 (openpyxl/reportlab): {str(e)}"
        logger.error(error_msg)
        return False, None, error_msg
    except Exception as e:
        error_msg = f"OpenPyXL转换失败: {str(e)}"
        logger.error(error_msg, exc_info=True) # Added exc_info=True for better debugging
        return False, None, error_msg

def convert_excel_to_pdf(input_path, output_pdf_path):
    """
    转换Excel到PDF的主函数，按优先级尝试不同方案
    
    Args:
        input_path: 输入的Excel文件路径(.xls/.xlsx)
        output_pdf_path: 输出的PDF文件路径
    
    Returns:
        tuple: (success: bool, actual_output_path: str or None, error_message: str or None)
    """
    logger.info(f"开始转换Excel到PDF: {input_path} -> {output_pdf_path}")
    
    # --- 方案1: LibreOffice (using the generic converter) ---
    output_dir_for_lo = os.path.dirname(output_pdf_path)
    
    # Ensure output directory exists for LO converter (lo_convert_to_pdf also does this, but good practice here too)
    if not os.path.isdir(output_dir_for_lo):
        try:
            os.makedirs(output_dir_for_lo, exist_ok=True)
        except OSError as e_mkdir:
            logger.error(f"创建LibreOffice输出目录失败 '{output_dir_for_lo}': {e_mkdir}")
            # This is a critical failure before calling LO, so we might return early.
            # However, lo_convert_to_pdf will also report an error if output_dir is not found.
            # Let's allow it to proceed and let lo_convert_to_pdf handle its pre-checks.
            pass

    # Call the generic LibreOffice converter.
    # lo_convert_to_pdf expects an output_dir, not a full output_path.
    # It returns the actual path of the created PDF (e.g., input_filename.pdf in output_dir)
    # and the original filename as created by LO.
    lo_success, lo_actual_pdf_path_or_error, lo_original_filename = lo_convert_to_pdf(input_path, output_dir_for_lo)
    
    if lo_success:
        # If the path LibreOffice used (lo_actual_pdf_path_or_error) is not the final desired output_pdf_path,
        # we need to move/rename it.
        if lo_actual_pdf_path_or_error != output_pdf_path:
            try:
                if os.path.exists(output_pdf_path):
                    logger.warning(f"目标PDF路径 {output_pdf_path} 已存在。将进行覆盖。")
                    os.remove(output_pdf_path)
                os.rename(lo_actual_pdf_path_or_error, output_pdf_path)
                logger.info(f"Excel到PDF（LibreOffice）：成功将 '{lo_actual_pdf_path_or_error}' 重命名/移动到 '{output_pdf_path}'")
                # The third element from lo_convert_to_pdf was lo_original_filename;
                # now we return the new basename if rename occurred.
                return True, output_pdf_path, os.path.basename(output_pdf_path) 
            except Exception as e_move:
                error_msg_move = f"Excel到PDF（LibreOffice）转换成功，但重命名/移动文件失败 从 '{lo_actual_pdf_path_or_error}' 到 '{output_pdf_path}': {e_move}"
                logger.error(error_msg_move)
                # Return success as True because PDF was created, but provide the path where it is, and the error message.
                return True, lo_actual_pdf_path_or_error, error_msg_move # Error message as the third element
        else:
            # File is already at the desired output_pdf_path
            logger.info(f"Excel到PDF（LibreOffice）成功，文件已在: {output_pdf_path}")
            return True, output_pdf_path, lo_original_filename # Original name from LO
    else:
        # lo_actual_pdf_path_or_error contains the error message from lo_convert_to_pdf
        logger.warning(f"LibreOffice (通用转换器) 转换Excel到PDF失败: {lo_actual_pdf_path_or_error}")
    
    # --- 方案2: OpenPyXL + ReportLab - 自定义方案，现在支持中文字体 ---
    logger.info("LibreOffice转换失败，尝试使用OpenPyXL+ReportLab方案。")
    openpyxl_success, openpyxl_result_path, openpyxl_error = convert_excel_to_pdf_openpyxl(input_path, output_pdf_path)
    if openpyxl_success:
        # convert_excel_to_pdf_openpyxl already saves to output_pdf_path, so no move needed here.
        # The third element it returns is None on success or an error message.
        return True, openpyxl_result_path, None 
    
    logger.warning(f"OpenPyXL+ReportLab转换也失败: {openpyxl_error}")
    
    # 所有方案都失败
    # Use the error from the last attempted primary method (LibreOffice) or the fallback (OpenPyXL)
    final_error_detail = openpyxl_error if openpyxl_error else lo_actual_pdf_path_or_error # lo_actual_pdf_path_or_error has LO error msg
    final_error_msg = f"所有Excel转PDF转换方案(LibreOffice, OpenPyXL+ReportLab)都失败。最后错误: {final_error_detail}"
    logger.error(final_error_msg)
    return False, None, final_error_msg 